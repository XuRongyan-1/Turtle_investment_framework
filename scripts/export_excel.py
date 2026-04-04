"""
Excel导出模块 - 将报告扫描结果导出为Excel表格
"""

import os
from pathlib import Path
from typing import List, Optional
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: pandas/openpyxl未安装，将使用CSV格式导出")

from scan_reports import ReportScanner, ReportData


class ReportExporter:
    """报告导出器"""

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.scanner = ReportScanner(output_dir)

    def scan_and_export(self, excel_file: Optional[str] = None) -> str:
        """扫描并导出到Excel"""
        # 扫描所有报告
        reports = self.scanner.scan_all()

        if not reports:
            print("未找到任何报告文件")
            return ""

        # 生成文件名
        if not excel_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = f"reports_summary_{timestamp}.xlsx"

        # 确保输出目录存在
        output_path = self.output_dir / excel_file

        if EXCEL_AVAILABLE:
            self._export_to_excel(reports, output_path)
        else:
            # 回退到CSV
            csv_path = output_path.with_suffix('.csv')
            self._export_to_csv(reports, csv_path)
            return str(csv_path)

        return str(output_path)

    def _export_to_excel(self, reports: List[ReportData], output_path: Path):
        """导出到Excel格式（带格式）"""
        # 准备数据
        data = []
        for r in reports:
            data.append({
                '股票代码': r.stock_code,
                '公司名称': r.company_name,
                '报告日期': r.report_date,

                # 四因子评分
                '因子1-盈利能力': r.factor1_profitability,
                '因子2-资本配置': r.factor2_capital_alloc,
                '因子3-安全边际': r.factor3_safety_margin,
                '因子4-护城河': r.factor4_moat,
                '综合评分': r.overall_score,

                # 价格数据
                '当前价格': r.current_price,
                '币种': r.price_currency,

                # 历史分位
                '历史分位(%)': r.hist_percentile,
                '中位数价格': r.median_price,

                # 折价/溢价率
                '相对中位数溢价率(%)': round(r.discount_premium_ratio, 2) if r.discount_premium_ratio >= 0 else 0,
                '相对中位数折价率(%)': abs(round(r.discount_premium_ratio, 2)) if r.discount_premium_ratio < 0 else 0,

                # 价格区间建议
                '增持上限': r.buy_threshold,
                '配置区间下限': r.accumulate_low,
                '配置区间上限': r.accumulate_high,
                '持有区间下限': r.hold_low,
                '持有区间上限': r.hold_high,
                '减仓下限': r.sell_threshold,

                # 风险评级
                '风险等级': r.risk_level,

                # 源文件
                '源文件': r.source_file
            })

        df = pd.DataFrame(data)

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "报告汇总"

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)

        # 设置列宽
        column_widths = {
            'A': 12,  # 股票代码
            'B': 25,  # 公司名称
            'C': 12,  # 报告日期
            'D': 12,  # 因子1
            'E': 12,  # 因子2
            'F': 12,  # 因子3
            'G': 12,  # 因子4
            'H': 12,  # 综合评分
            'I': 12,  # 当前价格
            'J': 8,   # 币种
            'K': 12,  # 历史分位
            'L': 12,  # 中位数价格
            'M': 15,  # 溢价率
            'N': 15,  # 折价率
            'O': 12,  # 增持上限
            'P': 14,  # 配置区间下限
            'Q': 14,  # 配置区间上限
            'R': 14,  # 持有区间下限
            'S': 14,  # 持有区间上限
            'T': 12,  # 减仓下限
            'U': 10,  # 风险等级
            'V': 50,  # 源文件
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置数据行样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 评分单元格颜色映射
        score_fill_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
        score_fill_medium = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
        score_fill_low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 为评分列添加颜色
            score_cols = [4, 5, 6, 7, 8]  # D-H列（因子1-4和综合评分）
            for col_idx in score_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    score = float(cell.value) if cell.value else 0
                    if score >= 8:
                        cell.fill = score_fill_high
                    elif score >= 6:
                        cell.fill = score_fill_medium
                    else:
                        cell.fill = score_fill_low
                except (ValueError, TypeError):
                    pass

            # 风险等级颜色
            risk_cell = ws.cell(row=row_idx, column=21)  # U列
            if risk_cell.value == "高":
                risk_cell.fill = score_fill_low
            elif risk_cell.value == "中":
                risk_cell.fill = score_fill_medium
            else:
                risk_cell.fill = score_fill_high

        # 冻结首行
        ws.freeze_panes = "A2"

        # 添加筛选
        ws.auto_filter.ref = ws.dimensions

        # 保存文件
        wb.save(output_path)
        print(f"Excel文件已保存: {output_path}")

        # 打印统计信息
        self._print_summary(reports)

    def _export_to_csv(self, reports: List[ReportData], output_path: Path):
        """导出到CSV格式（备用）"""
        data = []
        for r in reports:
            data.append({
                '股票代码': r.stock_code,
                '公司名称': r.company_name,
                '综合评分': r.overall_score,
                '因子1': r.factor1_profitability,
                '因子2': r.factor2_capital_alloc,
                '因子3': r.factor3_safety_margin,
                '因子4': r.factor4_moat,
                '当前价格': r.current_price,
                '历史分位': r.hist_percentile,
                '溢价/折价率': r.discount_premium_ratio,
                '风险等级': r.risk_level
            })

        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"CSV文件已保存: {output_path}")

    def _print_summary(self, reports: List[ReportData]):
        """打印汇总统计"""
        print("\n" + "="*80)
        print("报告汇总统计")
        print("="*80)

        print(f"\n总计报告数: {len(reports)}")

        # 按市场分类
        markets = {}
        for r in reports:
            market = r.stock_code.split('.')[-1] if '.' in r.stock_code else '未知'
            markets[market] = markets.get(market, 0) + 1

        print("\n按市场分布:")
        for market, count in sorted(markets.items()):
            print(f"  {market}: {count} 份")

        # 评分分布
        scores = [r.overall_score for r in reports if r.overall_score > 0]
        if scores:
            print(f"\n评分统计:")
            print(f"  平均分: {sum(scores)/len(scores):.2f}")
            print(f"  最高分: {max(scores):.2f}")
            print(f"  最低分: {min(scores):.2f}")
            print(f"  高分股(≥8): {len([s for s in scores if s >= 8])} 只")
            print(f"  中分股(6-8): {len([s for s in scores if 6 <= s < 8])} 只")
            print(f"  低分股(<6): {len([s for s in scores if s < 6])} 只")

        # 风险分布
        risk_dist = {}
        for r in reports:
            risk_dist[r.risk_level] = risk_dist.get(r.risk_level, 0) + 1

        print(f"\n风险等级分布:")
        for level in ['低', '中', '高']:
            if level in risk_dist:
                print(f"  {level}风险: {risk_dist[level]} 只")


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description='扫描报告并导出Excel汇总')
    parser.add_argument('--output', '-o', default=None, help='输出文件名')
    parser.add_argument('--dir', '-d', default='output', help='报告目录')

    args = parser.parse_args()

    exporter = ReportExporter(args.dir)
    output_file = exporter.scan_and_export(args.output)

    if output_file:
        print(f"\n导出完成: {output_file}")
    else:
        print("\n导出失败")


if __name__ == "__main__":
    main()
